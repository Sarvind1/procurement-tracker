import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
import shutil
from openpyxl.worksheet.table import Table, TableStyleInfo
import ast

warnings.filterwarnings('ignore', category=FutureWarning)

# # universal_path = r"C:\Users\SupplierCommunicatio\OneDrive - Razor HQ GmbH & Co. KG\Razor - Chetan_Locale\Procurement Trackers"
# universal_path = r"C:\Users\ChetanPaliwal\OneDrive - Razor HQ GmbH & Co. KG\Razor - Procurement Trackers"

def main(df1, list_pull, universal_path):
    # df1 = pd.read_csv('./exports/otif_dash.csv')

    exclude_list = [
        'PO381255','PO382140','PO382141','PO382115','PO382116','PO382117','PO382118','PO382121','PO382122','PO382123','PO382124','PO382125',
        'PO382126','PO382127','PO382128','PO381292','PO381293','PO378591','PO379182','PO381986','PO381811','PO381828','PO381999','PO378931',
        'PO381323','PO381258','PO381809','PO380821','PO381332','PO381815','PO381818','PO378152','PO379565','PO379633','PO379636','PO379644','PO380653',
        'PO380773','PO380783','PO380810','PO381322','PO381326','PO381315','PO381316','PO381320','PO381329','PO381333','PO380566','PO381277','PO381287',
        'PO380676','PO381801','PO381025','PO381183','PO381212','PO381261','PO381044','PO378800','PO381814','PO379569','PO378470','PO381817','PO381062','PO378839',
        'PO378473','PO381037','PO381038','PO381039','PO377215','PO378655','PO377410','PO378462','PO378526','PO378653','PO378654','PO378657','PO378719','PO378761','PO379136',
        'PO379578','PO379580','PO379582','PO379590','PO379599','PO379604','PO379608','PO379628','PO379637','PO379643','PO379647','PO379648','PO379649','PO379650','PO379651',
        'PO379652','PO379655','PO379656','PO379665','PO379671','PO379679','PO379698','PO379699','PO379722','PO379727','PO379732','PO379733','PO379737','PO379818','PO380666','PO380672',
        'PO380675','PO380682','PO380694','PO380696','PO380698','PO380717','PO380724','PO380727','PO380731','PO380751','PO380771','PO380789','PO380790','PO380798','PO380799','PO380801','PO380816','PO380823','PO380978','PO380985','PO381017','PO381018','PO381020','PO381027','PO381029','PO381032','PO381034','PO381035','PO381043','PO381045','PO381054','PO381055','PO381059','PO381060','PO381153','PO381155','PO381161','PO381177','PO381178','PO381180','PO381193','PO381199','PO381201','PO381203','PO381207','PO381213','PO381214','PO381216','PO381220','PO381221','PO381224','PO381228','PO381230','PO381232','PO381233','PO381236','PO381247','PO381249','PO381250','PO381251','PO381252','PO381254','PO381262','PO381266','PO381270','PO381310','PO381318','PO381319','PO381321','PO381324','PO381338','PO381349','PO381350','PO381351','PO381808','PO379728','PO374312','PO381264','PO379704','PO379705','PO378594','PO378722','PO381265','PO381257','PO381245','PO378502','PO381267','PO381179','PO381187','PO381200','PO381172','PO381242','PO381215','PO378715','PO377310','PO377316','PO378656','PO378658','PO379620','PO378500','PO378516','PO378586','PO378810','PO381191','PO381194','PO381246','PO381269','PO381330','PO381206','PO381210','PO381337','PO378816','PO378817','PO379585','PO380989','PO380723','PO380980','PO378525','PO380957','PO380994','PO377353','PO379610','PO379707','PO379740','PO381301','PO378507','PO378652','PO378704','PO378777','PO379562','PO380996','PO380971','PO380578','PO380657','PO379600','PO372914','PO379623','PO375299','PO378527','PO379581','PO379584','PO379586','PO380346','PO380468','PO380611','PO380625','PO380725','PO380792','PO380807','PO380965','PO378629','PO379646','PO380448','PO380459','PO380497','PO380550','PO380596','PO380614','PO380808','PO375331','PO375337','PO377442','PO378576','PO378581','PO378582','PO379579','PO379588','PO379702','PO380707','PO380730','PO380756','PO380761','PO380829','PO380952','PO381196','PO381354','PO378464','PO378474','PO378727','PO381256','PO381186','PO381229','PO379630','PO380689','PO380769','PO381154','PO381174','PO381185','PO381195','PO381209','PO381211','PO381227','PO381239','PO381244','PO381311','PO380684','PO380744','PO380813','PO381208','PO378615','PO378729','PO375288','PO381226','PO380794','PO381218','PO380699','PO377309','PO380741','PO380781','PO381276','PO377214','PO377261','PO378469','PO378555','PO379594','PO379662','PO379666','PO379677','PO379695','PO379697','PO379711','PO379714','PO379724','PO379736','PO379739','PO375327','PO375350','PO375361','PO375362','PO375444','PO375446','PO375466','PO377262','PO378448','PO378589','PO378571','PO378746','PO380702','PO381205','PO381222','PO378760','PO380784','PO381192','PO380706','PO380809','PO378758','PO380695','PO380760','PO378650','PO378651','PO379603','PO377311','PO378736','PO379715','PO377184','PO378484','PO380655','PO380749','PO380787','PO381231','PO381234','PO381353','PO378466','PO378475','PO378476','PO378489','PO378660','PO378661','PO378837','PO379602','PO379682','PO380776','PO377213','PO374370','PO377367','PO378478','PO377382','PO377199','PO377437','PO378556','PO378787','PO379612','PO378583','PO378585','PO378588','PO380667','PO377355','PO378734','PO377258','PO378435','PO378498','PO377399','PO378695','PO379641','PO379678','PO378844','PO378662','PO371096','PO378480','PO378499','PO378522','PO378562','PO378713','PO378789','PO378552','PO378732','PO378748','PO378773','PO377272','PO377385','PO377218','PO378647','PO378808','PO379561','PO377305','PO378436','PO378449','PO378494','PO378495','PO378517','PO378620','PO378621','PO379571','PO379591','PO379615','PO377429','PO377591','PO378465','PO378467','PO378573','PO378575','PO378602','PO378614','PO378637','PO378788','PO378824','PO378929','PO379611','PO379639','PO379725','PO378765','PO378687','PO378751','PO378851','PO379573','PO379575','PO378043','PO382005','PO382142','PO382143','PO382144','PO382145','PO382146','PO382147','PO382148'
    ]

    df1 = df1[~df1["document number"].isin(exclude_list)].copy()


    df1['id'] = df1['name'].str.split(" ").str[0]

    main_path = os.path.join(universal_path, "Compliance")
    today = pd.to_datetime(datetime.today().date())

    df1['prd'] = pd.to_datetime(df1['prd'], errors='coerce')
    df1['quality control date'] = pd.to_datetime(df1['quality control date'], errors='coerce')

    df1['po_razin_id'] = (
        df1['document number'].astype(str) + df1['item'].astype(str) + df1['line id'].astype(str)
    )

    # df_x = df1.copy()

    df1 = df1[~df1['quality control status'].str.lower().str.contains("not applicable", na=False)]


    def check_all_conditions(group):
        cond1 = (group['hs | sign-off shipment booking im line'] == "Yes").any()
        prd_check = group['prd'].apply(
            lambda x: False if pd.isna(x) or x == "" else (0 <= (x - today).days <= 10)
        )
        cond2 = prd_check.any()
        cond3 = group['quality control status'].str.lower().str.contains("awaiting release", na=False).all()
        cond4 = group['quality control date'].apply(lambda x: True if (pd.isna(x) or x == "") else False).all()
        return "Yes" if (cond1 and cond2 and cond3 and cond4) else "No"

    df_qi_email = df1.groupby("po_razin_id").apply(check_all_conditions).reset_index()
    df_qi_email.columns = ['po_razin_id', 'all_conditions_met']

    valid_pos = df_qi_email[df_qi_email['all_conditions_met'] == "Yes"]['po_razin_id']

    # list_pull = pd.read_csv(f'{main_path}\\QI Email Scheduler\\quality_inspection_compliance_list_pull.csv')

    final_cols = [
        'Title', 'batch_id', 'document number', 'automatedsent_x003f_', 'emailscheduledfor', 'sendearly',
        'prd', 'razin', 'brand', 'vendorid', 'vendorname', 'vendoremail',
        'cmemail', 'smemail', 'update_type', 'ID', 'compliancestatus', 'QIResult', 'Subset'
    ]
    for col in final_cols:
        if col not in list_pull.columns:
            list_pull[col] = ""

    list_pull = list_pull[final_cols]

    list_pull = list_pull.rename(columns={
        "Title": "po_razin_id", "automatedsent_x003f_": "sent already?",
        "emailscheduledfor": "email scheduled for", 'sendearly': 'send early?',
        'vendorid': 'vendor id', 'vendorname': 'vendor name',
        'vendoremail': 'vendor email', 'cmemail': 'cm email', 'smemail': 'sm email', 'compliancestatus': 'compliance status',
        'QIResult': 'quality control status',
        'ID': 'list_id',
        'Subset': 'subset'
    })

    def extract_value(x):
        if pd.isna(x):
            return None
        try:
            return ast.literal_eval(x)["Value"]
        except Exception:
            return x
    list_pull["send early?"] = list_pull["send early?"].apply(extract_value)

    early_docs = list_pull.loc[list_pull['send early?'] == 'Yes', 'po_razin_id'].unique()
    # print(early_docs)
    valid_pos = pd.concat([pd.Series(valid_pos), pd.Series(early_docs)]).unique()

    df_qi_email = df1[df1['po_razin_id'].isin(valid_pos)][
        ['po_razin_id', 'document number', 'item', 'line id', 'id', 'associated brands']
    ].sort_values(by=["po_razin_id"])
    df_qi_email['date sent'] = today
    df_qi_email['prd change so sent again'] = "No"

    sent_csv_path = os.path.join(main_path, "QI Email Scheduler", "already_sent.csv")

    if os.path.exists(sent_csv_path):
        sent_df = pd.read_csv(sent_csv_path, dtype=str)
    else:
        sent_df = pd.DataFrame(columns=['po_razin_id', 'document number', 'item', 'line id', 'date sent'])

    df_qi_email[['po_razin_id', 'document number', 'item', 'line id', 'prd change so sent again']] = df_qi_email[['po_razin_id', 'document number', 'item', 'line id', 'prd change so sent again']].astype(str)
    sent_df[['po_razin_id', 'document number', 'item', 'line id', 'prd change so sent again']] = sent_df[['po_razin_id', 'document number', 'item', 'line id', 'prd change so sent again']].astype(str)

    sent_df_pos = sent_df[sent_df['prd change so sent again'] == "No"]['document number']
    subset = df1[df1['document number'].isin(sent_df_pos)]
    mismatch_rows = subset[subset['prd'] != subset['first prd']]

    sent_df_qi_email = mismatch_rows.groupby("document number").apply(check_all_conditions)
    sent_df_qi_email = sent_df_qi_email.reset_index(name="all_conditions_met")
    sent_df_valid_pos = sent_df_qi_email[sent_df_qi_email['all_conditions_met'] == "Yes"]['document number']
    sent_df_qi_email = subset.copy()
    sent_df_qi_email = sent_df_qi_email[sent_df_qi_email['document number'].isin(sent_df_valid_pos)][['po_razin_id', 'document number', 'item', 'line id', 'id', 'associated brands']].sort_values(by=["document number", "line id"])
    sent_df_qi_email['date sent'] = today
    sent_df_qi_email['prd change so sent again'] = "Yes"

    df_qi_email = pd.concat([df_qi_email, sent_df_qi_email], ignore_index=True)

    df_qi_email['key'] = df_qi_email['document number'].astype(str)  + '|' + df_qi_email['item'].astype(str)  + '|' + df_qi_email['line id'].astype(str)  + '|' + df_qi_email['prd change so sent again'].astype(str) 
    sent_df['key'] = sent_df['document number'].astype(str)  + '|' + sent_df['item'].astype(str)  + '|' + sent_df['line id'].astype(str)  + '|' + sent_df['prd change so sent again'].astype(str) 

    new_lines_df = df_qi_email[~df_qi_email['key'].isin(sent_df['key'])].copy()

    df_qi_email.drop(columns='key', inplace=True)
    sent_df.drop(columns='key', inplace=True)
    new_lines_df.drop(columns='key', inplace=True)


    excel_template_path = os.path.join(main_path, "QI Email Scheduler", "QI Booking Form.xlsx")
    output_folder = os.path.join(main_path, "QI Email Scheduler", "attachments")
    os.makedirs(output_folder, exist_ok=True)

    grouped_by_doc = new_lines_df.groupby('document number', sort=False)

    for doc_number, group_df in grouped_by_doc:
        new_file_path = os.path.join(output_folder, f"{doc_number}_QI_Booking_Form.xlsx")
        shutil.copy(excel_template_path, new_file_path)
        wb = load_workbook(new_file_path)
        base_sheet_name = "Package Information"
        if base_sheet_name not in wb.sheetnames:
            raise Exception(f"Sheet '{base_sheet_name}' not found in template.")

        for item in group_df['item'].unique():
            source_sheet = wb[base_sheet_name]
            new_sheet = wb.copy_worksheet(source_sheet)
            new_sheet.title = f"{doc_number} - {item}"

        if base_sheet_name in wb.sheetnames:
            std = wb[base_sheet_name]
            wb.remove(std)
        wb.save(new_file_path)

    base_dir = "/Shared Documents/Chetan_Locale/Procurement Trackers/Compliance/QI Email Scheduler/attachments/"
    new_lines_df['directory'] = base_dir + new_lines_df['document number'].astype(str) + "_QI_Booking_Form.xlsx"

    log_df = new_lines_df.drop_duplicates(subset='document number', keep='first')[['document number', 'directory', 'id', 'associated brands']]

    vendor_info_path = os.path.join(universal_path, "Compliance", "QI Email Scheduler", "vendor_email_map.xlsx")
    vendor_df = pd.read_excel(vendor_info_path)

    log_df['id'] = pd.to_numeric(log_df['id'], errors='coerce').fillna(0).astype(int)
    vendor_df['id'] = pd.to_numeric(vendor_df['id'], errors='coerce').fillna(0).astype(int)

    merged_df = log_df.merge(vendor_df.drop_duplicates(subset='id', keep='first')[['id', 'name', 'email', 'cm email', 'sm email']], on='id', how='left')

    today_str = datetime.now().strftime("%m_%d_%Y")
    log_file_path = os.path.join(main_path, "QI Email Scheduler", "send logs", f"Log_{today_str}.xlsx")
    os.makedirs(os.path.dirname(log_file_path), exist_ok=True)

    if not merged_df.empty:
        merged_df.to_excel(log_file_path, index=False)
        wb = load_workbook(log_file_path)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column
        table_range = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="Log", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                            showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(log_file_path)

        df_concat = pd.concat([sent_df, new_lines_df], ignore_index=True, sort=False)
        df_concat['date sent'] = pd.to_datetime(df_concat['date sent'], errors='coerce').dt.date
        out_cols = ['po_razin_id', 'document number', 'item', 'line id', 'date sent', 'prd change so sent again']

        for c in out_cols:
            if c not in df_concat.columns:
                df_concat[c] = ""
        df_concat[out_cols].to_csv(sent_csv_path, index=True)
        print("Automation Complete for Compliance QI Email Scheduler")
    else:
        print("No new emails to send for Compliance QI Email Scheduler")