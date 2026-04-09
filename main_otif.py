from manipulate_otif import manipulate
from fetch_redshift_func import run_redshift_query
from init_pull_push import main
from fetch_csv import main as main_fetch_csv
from datetime import datetime, timezone
from send_update_slack import send_simple_slack_message
import pandas as pd
from openpyxl import load_workbook

start_ts = datetime.now(timezone.utc)
print(f"\nSTART TIME (UTC): {start_ts}\n")

list_name = "OTIF Team Trackers"

df1 = main_fetch_csv(list_name)

def load_query(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()

def read_excel_table(file_path, sheet_name, table_name):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    table = ws.tables[table_name]
    data = ws[table.ref]
    rows = [[cell.value for cell in row] for row in data]
    return pd.DataFrame(rows[1:], columns=rows[0])

query_1 = load_query("Queries/otif_dash.txt")
query_2 = load_query("Queries/otif_tracker.txt")
query_3 = load_query("Queries/rgbit_netsuite.txt")
query_4 = load_query("Queries/shipment_telex.txt")

# # universal_path = r"C:\Users\SupplierCommunicatio\OneDrive - Razor HQ GmbH & Co. KG\Tech & Analytics - OTIF"
# universal_path = r"C:\Users\ChetanPaliwal\OneDrive - Razor HQ GmbH & Co. KG\Tech & Analytics - OTIF"
# df2 = read_excel_table(rf"{universal_path}\OTIF File for DWH Import V4.xlsx", "Data", "OTIF_data")

df2 = run_redshift_query(query_1, "otif_dash")
df3 = run_redshift_query(query_2, "otif_tracker")
df4 = run_redshift_query(query_3, "rgbit_netsuite")
df5 = run_redshift_query(query_4, "shipment_telex")


df2_manipulated = manipulate(df2, df3, df4, df5) ## process specific result

main(list_name, df1, df2_manipulated)

end_ts = datetime.now(timezone.utc)
minutes_taken = round((end_ts - start_ts).total_seconds() / 60, 2)
print(f"\nEND TIME (UTC): {end_ts}")
print(f"TOTAL TIME TAKEN: {minutes_taken} minutes\n")

send_simple_slack_message("OTIF List Pull-Push Successful")